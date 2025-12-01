"""
Utilities for exporting data to Excel and PDF formats
"""
from datetime import datetime
from io import BytesIO
from typing import List, Dict, Any
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.platypus.flowables import Image as RLImage
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT

from core.constants import CURRENCY_SYMBOLS


def format_currency(amount: float, currency: str = "DOP") -> str:
    """Format currency with symbol"""
    symbol = CURRENCY_SYMBOLS.get(currency, "$")
    return f"{symbol} {amount:,.2f}"


def create_excel_comparison(
    data: Dict[str, Any],
    title: str,
    year: int,
    company_name: str = "OnQuota"
) -> BytesIO:
    """
    Create an Excel file with comparison data

    Args:
        data: Dictionary containing 'monthly_data' and 'summary'
        title: Report title (e.g., "Comparación de Gastos")
        year: Year of the report
        company_name: Name of the company

    Returns:
        BytesIO: Excel file in memory
    """
    wb = Workbook()
    ws = wb.active
    ws.title = f"{title} {year}"

    # Header styling
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    subheader_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    subheader_font = Font(bold=True, size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Title
    ws.merge_cells('A1:F1')
    title_cell = ws['A1']
    title_cell.value = f"{company_name} - {title}"
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal='center')
    title_cell.fill = header_fill
    title_cell.font = Font(bold=True, size=14, color="FFFFFF")

    # Subtitle
    ws.merge_cells('A2:F2')
    subtitle_cell = ws['A2']
    subtitle_cell.value = f"Año {year} - Generado el {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    subtitle_cell.alignment = Alignment(horizontal='center')
    subtitle_cell.font = Font(size=10, italic=True)

    # Empty row
    ws.append([])

    # Summary section
    row_num = 4
    ws[f'A{row_num}'] = "RESUMEN EJECUTIVO"
    ws[f'A{row_num}'].font = subheader_font
    ws[f'A{row_num}'].fill = subheader_fill
    ws.merge_cells(f'A{row_num}:F{row_num}')

    summary = data.get('summary', {})

    row_num += 1
    summary_data = [
        ["Total Año Actual:", format_currency(summary.get('total_actual', 0))],
        ["Total Año Anterior:", format_currency(summary.get('total_previous', 0))],
        ["Cambio Porcentual:", f"{summary.get('percent_change', 0):.2f}%"],
    ]

    # Add type-specific summary fields
    if 'average_monthly' in summary:
        summary_data.append(["Promedio Mensual:", format_currency(summary['average_monthly'])])
    if 'total_quotes' in summary:
        summary_data.append(["Total Cotizaciones:", str(summary['total_quotes'])])
    if 'acceptance_rate' in summary:
        summary_data.append(["Tasa de Aceptación:", f"{summary['acceptance_rate']}%"])
    if 'average_ticket' in summary:
        summary_data.append(["Ticket Promedio:", format_currency(summary['average_ticket'])])

    for label, value in summary_data:
        ws[f'A{row_num}'] = label
        ws[f'B{row_num}'] = value
        ws[f'A{row_num}'].font = Font(bold=True)
        row_num += 1

    # Empty rows
    ws.append([])
    ws.append([])
    row_num += 2

    # Monthly data section
    ws[f'A{row_num}'] = "DATOS MENSUALES"
    ws[f'A{row_num}'].font = subheader_font
    ws[f'A{row_num}'].fill = subheader_fill
    ws.merge_cells(f'A{row_num}:F{row_num}')

    row_num += 1

    # Column headers
    monthly_data = data.get('monthly_data', [])
    if monthly_data and len(monthly_data) > 0:
        # Determine columns based on data structure
        first_item = monthly_data[0]
        headers = ['Mes', 'Año Actual', 'Año Anterior', 'Diferencia', '% Cambio']

        if 'count' in first_item:
            headers.append('Cantidad Actual')
        if 'accepted_count' in first_item:
            headers.append('Aceptadas')

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border

        row_num += 1

        # Data rows
        for item in monthly_data:
            actual = item.get('actual', 0)
            previous = item.get('previous', 0)
            difference = actual - previous
            pct_change = ((actual - previous) / previous * 100) if previous > 0 else 0

            row_data = [
                item.get('month', ''),
                actual,
                previous,
                difference,
                f"{pct_change:.2f}%"
            ]

            if 'count' in item:
                row_data.append(item['count'])
            if 'accepted_count' in item:
                row_data.append(item['accepted_count'])

            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = value
                cell.border = border

                # Format numbers as currency (columns 2, 3, 4)
                if col_num in [2, 3, 4] and isinstance(value, (int, float)):
                    cell.number_format = '#,##0.00'

                # Center alignment for all except first column
                if col_num > 1:
                    cell.alignment = Alignment(horizontal='center')

            row_num += 1

    # Adjust column widths
    for col in range(1, 8):
        ws.column_dimensions[get_column_letter(col)].width = 18

    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return output


def create_pdf_comparison(
    data: Dict[str, Any],
    title: str,
    year: int,
    company_name: str = "OnQuota"
) -> BytesIO:
    """
    Create a PDF file with comparison data

    Args:
        data: Dictionary containing 'monthly_data' and 'summary'
        title: Report title
        year: Year of the report
        company_name: Name of the company

    Returns:
        BytesIO: PDF file in memory
    """
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    elements = []

    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#366092'),
        spaceAfter=30,
        alignment=TA_CENTER
    )

    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=14,
        textColor=colors.HexColor('#366092'),
        spaceAfter=12,
        spaceBefore=12
    )

    # Title
    elements.append(Paragraph(f"{company_name}", title_style))
    elements.append(Paragraph(f"{title} - Año {year}", title_style))
    elements.append(Paragraph(
        f"Generado el {datetime.now().strftime('%d de %B de %Y a las %H:%M')}",
        styles['Normal']
    ))
    elements.append(Spacer(1, 20))

    # Summary section
    elements.append(Paragraph("Resumen Ejecutivo", heading_style))

    summary = data.get('summary', {})
    summary_data = [
        ['Métrica', 'Valor'],
        ['Total Año Actual', format_currency(summary.get('total_actual', 0))],
        ['Total Año Anterior', format_currency(summary.get('total_previous', 0))],
        ['Cambio Porcentual', f"{summary.get('percent_change', 0):.2f}%"],
    ]

    # Add type-specific fields
    if 'average_monthly' in summary:
        summary_data.append(['Promedio Mensual', format_currency(summary['average_monthly'])])
    if 'total_quotes' in summary:
        summary_data.append(['Total Cotizaciones', str(summary['total_quotes'])])
    if 'acceptance_rate' in summary:
        summary_data.append(['Tasa de Aceptación', f"{summary['acceptance_rate']}%"])
    if 'average_ticket' in summary:
        summary_data.append(['Ticket Promedio', format_currency(summary['average_ticket'])])

    if summary.get('max_month'):
        summary_data.append([
            'Mejor Mes',
            f"{summary['max_month']['name']} - {format_currency(summary['max_month']['amount'])}"
        ])

    if summary.get('min_month'):
        summary_data.append([
            'Mes con Menor Gasto',
            f"{summary['min_month']['name']} - {format_currency(summary['min_month']['amount'])}"
        ])

    summary_table = Table(summary_data, colWidths=[3*inch, 3*inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTNAME', (0, 1), (0, -1), 'Helvetica-Bold'),
    ]))

    elements.append(summary_table)
    elements.append(Spacer(1, 30))

    # Monthly data section
    elements.append(Paragraph("Datos Mensuales Detallados", heading_style))

    monthly_data = data.get('monthly_data', [])
    if monthly_data and len(monthly_data) > 0:
        first_item = monthly_data[0]

        # Determine columns
        table_headers = ['Mes', 'Actual', 'Anterior', 'Diferencia', '% Cambio']
        col_widths = [1*inch, 1.2*inch, 1.2*inch, 1.2*inch, 1*inch]

        if 'count' in first_item:
            table_headers.append('Cantidad')
            col_widths.append(0.8*inch)
        if 'accepted_count' in first_item:
            table_headers.append('Aceptadas')
            col_widths.append(0.8*inch)

        table_data = [table_headers]

        for item in monthly_data:
            actual = item.get('actual', 0)
            previous = item.get('previous', 0)
            difference = actual - previous
            pct_change = ((actual - previous) / previous * 100) if previous > 0 else 0

            row = [
                item.get('month', ''),
                format_currency(actual),
                format_currency(previous),
                format_currency(difference),
                f"{pct_change:.1f}%"
            ]

            if 'count' in item:
                row.append(str(item['count']))
            if 'accepted_count' in item:
                row.append(str(item['accepted_count']))

            table_data.append(row)

        monthly_table = Table(table_data, colWidths=col_widths)
        monthly_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
        ]))

        elements.append(monthly_table)

    # Build PDF
    doc.build(elements)
    buffer.seek(0)

    return buffer
