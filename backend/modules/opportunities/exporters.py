"""
Opportunity Export Service
Handles exporting opportunities to various formats (Excel, CSV, etc.)
"""
import os
from datetime import datetime
from typing import List
from decimal import Decimal
from uuid import uuid4

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from models.opportunity import Opportunity, OpportunityStage
from core.logging import get_logger

logger = get_logger(__name__)


class OpportunityExporter:
    """
    Export opportunities to Excel format with multiple analysis sheets

    Features:
    - Multi-sheet workbook
    - Professional formatting
    - Summary statistics
    - Stage analysis
    - Sales rep performance
    """

    def __init__(self, export_dir: str = "/Users/josegomez/Documents/Code/SaaS/07 - OnQuota/backend/exports"):
        """
        Initialize exporter

        Args:
            export_dir: Directory to save exported files
        """
        self.export_dir = export_dir
        os.makedirs(export_dir, exist_ok=True)

    async def export_to_excel(
        self,
        opportunities: List[Opportunity],
        filename: str = None
    ) -> str:
        """
        Export opportunities to Excel with multiple sheets

        Sheets:
        1. All Opportunities - Complete list with all fields
        2. Summary by Stage - Aggregated statistics per stage
        3. Summary by Sales Rep - Performance by sales representative
        4. Win Rate Analysis - Win/loss analysis

        Args:
            opportunities: List of Opportunity objects to export
            filename: Optional filename (auto-generated if not provided)

        Returns:
            str: Full filepath to generated Excel file
        """
        try:
            # Generate filename if not provided
            if not filename:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"opportunities_{timestamp}.xlsx"

            # Ensure .xlsx extension
            if not filename.endswith('.xlsx'):
                filename += '.xlsx'

            filepath = os.path.join(self.export_dir, filename)

            # Create workbook
            wb = Workbook()

            # Remove default sheet
            wb.remove(wb.active)

            # Create sheets
            self._create_opportunities_sheet(wb, opportunities)
            self._create_stage_summary_sheet(wb, opportunities)
            self._create_sales_rep_summary_sheet(wb, opportunities)
            self._create_win_rate_sheet(wb, opportunities)

            # Save workbook
            wb.save(filepath)

            logger.info(f"Exported {len(opportunities)} opportunities to {filepath}")

            return filepath

        except Exception as e:
            logger.error(f"Error exporting opportunities to Excel: {str(e)}")
            raise

    def _create_opportunities_sheet(self, wb: Workbook, opportunities: List[Opportunity]):
        """Create sheet with all opportunities listed"""
        ws = wb.create_sheet("All Opportunities")

        # Define headers
        headers = [
            "ID", "Name", "Client", "Sales Rep", "Stage", "Estimated Value",
            "Currency", "Probability (%)", "Weighted Value", "Expected Close",
            "Actual Close", "Loss Reason", "Created At", "Updated At"
        ]

        # Write headers with styling
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Write data rows
        for row_num, opp in enumerate(opportunities, 2):
            ws.cell(row=row_num, column=1, value=str(opp.id))
            ws.cell(row=row_num, column=2, value=opp.name)
            ws.cell(row=row_num, column=3, value=opp.client.name if opp.client else "N/A")
            ws.cell(row=row_num, column=4, value=opp.sales_rep.full_name if opp.sales_rep else "N/A")
            ws.cell(row=row_num, column=5, value=opp.stage.value)
            ws.cell(row=row_num, column=6, value=float(opp.estimated_value))
            ws.cell(row=row_num, column=7, value=opp.currency)
            ws.cell(row=row_num, column=8, value=float(opp.probability))
            ws.cell(row=row_num, column=9, value=opp.weighted_value)
            ws.cell(row=row_num, column=10, value=opp.expected_close_date)
            ws.cell(row=row_num, column=11, value=opp.actual_close_date)
            ws.cell(row=row_num, column=12, value=opp.loss_reason or "")
            ws.cell(row=row_num, column=13, value=opp.created_at)
            ws.cell(row=row_num, column=14, value=opp.updated_at)

        # Format currency columns
        for row in range(2, len(opportunities) + 2):
            ws.cell(row=row, column=6).number_format = '#,##0.00'
            ws.cell(row=row, column=9).number_format = '#,##0.00'

        # Auto-adjust column widths
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15

        # Make name column wider
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 25

        # Freeze header row
        ws.freeze_panes = "A2"

    def _create_stage_summary_sheet(self, wb: Workbook, opportunities: List[Opportunity]):
        """Create sheet with summary by stage"""
        ws = wb.create_sheet("Summary by Stage")

        # Headers
        headers = ["Stage", "Count", "Total Value", "Weighted Value", "Avg Probability", "Avg Value"]
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Calculate statistics by stage
        stage_stats = {}
        for stage in OpportunityStage:
            stage_opps = [opp for opp in opportunities if opp.stage == stage]
            if stage_opps:
                count = len(stage_opps)
                total_value = sum(float(opp.estimated_value) for opp in stage_opps)
                weighted_value = sum(opp.weighted_value for opp in stage_opps)
                avg_probability = sum(float(opp.probability) for opp in stage_opps) / count
                avg_value = total_value / count

                stage_stats[stage.value] = {
                    "count": count,
                    "total_value": total_value,
                    "weighted_value": weighted_value,
                    "avg_probability": avg_probability,
                    "avg_value": avg_value
                }

        # Write data
        row_num = 2
        for stage_name, stats in stage_stats.items():
            ws.cell(row=row_num, column=1, value=stage_name)
            ws.cell(row=row_num, column=2, value=stats["count"])
            ws.cell(row=row_num, column=3, value=stats["total_value"])
            ws.cell(row=row_num, column=4, value=stats["weighted_value"])
            ws.cell(row=row_num, column=5, value=stats["avg_probability"])
            ws.cell(row=row_num, column=6, value=stats["avg_value"])

            # Format currency
            ws.cell(row=row_num, column=3).number_format = '#,##0.00'
            ws.cell(row=row_num, column=4).number_format = '#,##0.00'
            ws.cell(row=row_num, column=5).number_format = '0.00'
            ws.cell(row=row_num, column=6).number_format = '#,##0.00'

            row_num += 1

        # Totals row
        total_row = row_num + 1
        ws.cell(row=total_row, column=1, value="TOTAL")
        ws.cell(row=total_row, column=1).font = Font(bold=True)
        ws.cell(row=total_row, column=2, value=len(opportunities))
        ws.cell(row=total_row, column=3, value=sum(float(opp.estimated_value) for opp in opportunities))
        ws.cell(row=total_row, column=4, value=sum(opp.weighted_value for opp in opportunities))

        # Format totals
        ws.cell(row=total_row, column=3).number_format = '#,##0.00'
        ws.cell(row=total_row, column=4).number_format = '#,##0.00'

        # Auto-adjust column widths
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 18

        ws.freeze_panes = "A2"

    def _create_sales_rep_summary_sheet(self, wb: Workbook, opportunities: List[Opportunity]):
        """Create sheet with summary by sales rep"""
        ws = wb.create_sheet("Summary by Sales Rep")

        # Headers
        headers = ["Sales Rep", "Total Opps", "Won", "Lost", "Active", "Win Rate %",
                   "Total Won Value", "Avg Deal Size", "Pipeline Value"]
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Calculate statistics by sales rep
        rep_stats = {}
        for opp in opportunities:
            rep_name = opp.sales_rep.full_name if opp.sales_rep else "Unassigned"

            if rep_name not in rep_stats:
                rep_stats[rep_name] = {
                    "total": 0,
                    "won": 0,
                    "lost": 0,
                    "active": 0,
                    "won_value": 0.0,
                    "pipeline_value": 0.0
                }

            rep_stats[rep_name]["total"] += 1

            if opp.stage == OpportunityStage.CLOSED_WON:
                rep_stats[rep_name]["won"] += 1
                rep_stats[rep_name]["won_value"] += float(opp.estimated_value)
            elif opp.stage == OpportunityStage.CLOSED_LOST:
                rep_stats[rep_name]["lost"] += 1
            else:
                rep_stats[rep_name]["active"] += 1
                rep_stats[rep_name]["pipeline_value"] += float(opp.estimated_value)

        # Write data
        row_num = 2
        for rep_name, stats in rep_stats.items():
            closed = stats["won"] + stats["lost"]
            win_rate = (stats["won"] / closed * 100) if closed > 0 else 0
            avg_deal = stats["won_value"] / stats["won"] if stats["won"] > 0 else 0

            ws.cell(row=row_num, column=1, value=rep_name)
            ws.cell(row=row_num, column=2, value=stats["total"])
            ws.cell(row=row_num, column=3, value=stats["won"])
            ws.cell(row=row_num, column=4, value=stats["lost"])
            ws.cell(row=row_num, column=5, value=stats["active"])
            ws.cell(row=row_num, column=6, value=win_rate)
            ws.cell(row=row_num, column=7, value=stats["won_value"])
            ws.cell(row=row_num, column=8, value=avg_deal)
            ws.cell(row=row_num, column=9, value=stats["pipeline_value"])

            # Format numbers
            ws.cell(row=row_num, column=6).number_format = '0.00'
            ws.cell(row=row_num, column=7).number_format = '#,##0.00'
            ws.cell(row=row_num, column=8).number_format = '#,##0.00'
            ws.cell(row=row_num, column=9).number_format = '#,##0.00'

            row_num += 1

        # Auto-adjust column widths
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 16

        ws.column_dimensions['A'].width = 25
        ws.freeze_panes = "A2"

    def _create_win_rate_sheet(self, wb: Workbook, opportunities: List[Opportunity]):
        """Create sheet with win rate analysis"""
        ws = wb.create_sheet("Win Rate Analysis")

        # Calculate win rate metrics
        closed_opps = [opp for opp in opportunities
                       if opp.stage in [OpportunityStage.CLOSED_WON, OpportunityStage.CLOSED_LOST]]
        won_opps = [opp for opp in closed_opps if opp.stage == OpportunityStage.CLOSED_WON]
        lost_opps = [opp for opp in closed_opps if opp.stage == OpportunityStage.CLOSED_LOST]

        total_closed = len(closed_opps)
        total_won = len(won_opps)
        total_lost = len(lost_opps)
        win_rate = (total_won / total_closed * 100) if total_closed > 0 else 0

        won_value = sum(float(opp.estimated_value) for opp in won_opps)
        lost_value = sum(float(opp.estimated_value) for opp in lost_opps)
        avg_won_value = won_value / total_won if total_won > 0 else 0
        avg_lost_value = lost_value / total_lost if total_lost > 0 else 0

        # Title
        ws.cell(row=1, column=1, value="Win Rate Analysis")
        ws.cell(row=1, column=1).font = Font(bold=True, size=14)

        # Metrics section
        row = 3
        metrics = [
            ("Total Closed Opportunities", total_closed),
            ("Won Opportunities", total_won),
            ("Lost Opportunities", total_lost),
            ("Win Rate", f"{win_rate:.2f}%"),
            ("", ""),
            ("Total Won Value", won_value),
            ("Total Lost Value", lost_value),
            ("Average Won Deal Size", avg_won_value),
            ("Average Lost Deal Size", avg_lost_value),
        ]

        for metric_name, metric_value in metrics:
            ws.cell(row=row, column=1, value=metric_name)
            ws.cell(row=row, column=1).font = Font(bold=True)

            if isinstance(metric_value, (int, float)) and metric_name != "":
                ws.cell(row=row, column=2, value=metric_value)
                if "Value" in metric_name or "Size" in metric_name:
                    ws.cell(row=row, column=2).number_format = '#,##0.00'
            else:
                ws.cell(row=row, column=2, value=metric_value)

            row += 1

        # Loss reasons section
        if lost_opps:
            row += 2
            ws.cell(row=row, column=1, value="Loss Reasons")
            ws.cell(row=row, column=1).font = Font(bold=True, size=12)
            row += 1

            # Count loss reasons
            loss_reasons = {}
            for opp in lost_opps:
                reason = opp.loss_reason or "Not specified"
                loss_reasons[reason] = loss_reasons.get(reason, 0) + 1

            # Headers
            ws.cell(row=row, column=1, value="Reason")
            ws.cell(row=row, column=2, value="Count")
            ws.cell(row=row, column=1).font = Font(bold=True)
            ws.cell(row=row, column=2).font = Font(bold=True)
            row += 1

            # Write loss reasons
            for reason, count in sorted(loss_reasons.items(), key=lambda x: x[1], reverse=True):
                ws.cell(row=row, column=1, value=reason)
                ws.cell(row=row, column=2, value=count)
                row += 1

        # Auto-adjust column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
