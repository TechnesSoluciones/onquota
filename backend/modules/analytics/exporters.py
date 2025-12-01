"""
Exporters for Analytics Results
Export analysis results to Excel and PDF formats with formatting
"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import PieChart, Reference
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import (
    SimpleDocTemplate,
    Table,
    TableStyle,
    Paragraph,
    Spacer,
    PageBreak,
    Image,
)
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from datetime import datetime
from typing import Dict, List, Optional
from pathlib import Path
import logging

from models.analysis import Analysis

logger = logging.getLogger(__name__)


class ExcelExporter:
    """Export analysis results to formatted Excel file"""

    # Color scheme for ABC categories
    COLORS = {
        "A": "00B050",  # Green
        "B": "FFC000",  # Yellow/Orange
        "C": "FF0000",  # Red
        "header": "002060",  # Dark Blue
        "summary": "DCE6F1",  # Light Blue
    }

    @staticmethod
    def export_analysis(analysis: Analysis, output_path: str) -> str:
        """
        Export complete analysis to Excel with multiple sheets

        Args:
            analysis: Analysis model instance with results
            output_path: Path where to save the Excel file

        Returns:
            Path to the created Excel file

        Raises:
            ValueError: If analysis is not completed or has no results
        """
        if not analysis.is_completed or not analysis.results:
            raise ValueError("Analysis must be completed with results to export")

        logger.info(f"Exporting analysis {analysis.id} to Excel: {output_path}")

        # Create workbook
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet

        results = analysis.results

        # 1. Summary Sheet
        ExcelExporter._create_summary_sheet(wb, analysis, results)

        # 2. ABC Analysis Sheets
        if "abc_analysis" in results:
            ExcelExporter._create_abc_sheet(
                wb, results, "by_product", "ABC - Products"
            )
            if "by_client" in results.get("abc_analysis", {}):
                ExcelExporter._create_abc_sheet(
                    wb, results, "by_client", "ABC - Clients"
                )

        # 3. Top Products Sheet
        if "top_products" in results:
            ExcelExporter._create_top_items_sheet(
                wb, results["top_products"], "Top Products"
            )

        # 4. Top Clients Sheet
        if "top_clients" in results:
            ExcelExporter._create_top_items_sheet(
                wb, results["top_clients"], "Top Clients"
            )

        # 5. Discount Analysis Sheet
        if "discount_analysis" in results and results["discount_analysis"]:
            ExcelExporter._create_discount_sheet(wb, results["discount_analysis"])

        # 6. Margin Analysis Sheet
        if "margin_analysis" in results and results["margin_analysis"]:
            ExcelExporter._create_margin_sheet(wb, results["margin_analysis"])

        # 7. Monthly Trends Sheet
        if "monthly_trends" in results and results["monthly_trends"]:
            ExcelExporter._create_trends_sheet(wb, results["monthly_trends"])

        # 8. Insights Sheet
        if "insights" in results:
            ExcelExporter._create_insights_sheet(wb, results["insights"])

        # Save workbook
        wb.save(output_path)
        logger.info(f"Excel export completed: {output_path}")

        return output_path

    @staticmethod
    def _create_summary_sheet(wb: Workbook, analysis: Analysis, results: Dict):
        """Create summary statistics sheet"""
        ws = wb.create_sheet("Summary", 0)

        # Title
        ws["A1"] = "Sales Analysis Summary"
        ws["A1"].font = Font(size=16, bold=True, color="FFFFFF")
        ws["A1"].fill = PatternFill(start_color=ExcelExporter.COLORS["header"], fill_type="solid")
        ws.merge_cells("A1:B1")

        # Analysis metadata
        row = 3
        metadata = [
            ("Analysis Name:", analysis.name),
            ("Description:", analysis.description or "N/A"),
            ("Generated:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            ("Rows Processed:", f"{analysis.row_count:,}"),
        ]

        for label, value in metadata:
            ws[f"A{row}"] = label
            ws[f"A{row}"].font = Font(bold=True)
            ws[f"B{row}"] = value
            row += 1

        # Summary statistics
        row += 2
        ws[f"A{row}"] = "Key Metrics"
        ws[f"A{row}"].font = Font(size=14, bold=True)
        ws[f"A{row}"].fill = PatternFill(start_color=ExcelExporter.COLORS["summary"], fill_type="solid")
        ws.merge_cells(f"A{row}:B{row}")

        row += 1
        summary = results.get("summary", {})
        metrics = [
            ("Total Sales:", f"${summary.get('total_sales', 0):,.2f}"),
            ("Average Sale:", f"${summary.get('avg_sale', 0):,.2f}"),
            ("Median Sale:", f"${summary.get('median_sale', 0):,.2f}"),
            ("Std Deviation:", f"${summary.get('std_dev', 0):,.2f}"),
            ("Min Sale:", f"${summary.get('min_sale', 0):,.2f}"),
            ("Max Sale:", f"${summary.get('max_sale', 0):,.2f}"),
        ]

        for label, value in metrics:
            ws[f"A{row}"] = label
            ws[f"A{row}"].font = Font(bold=True)
            ws[f"B{row}"] = value
            ws[f"B{row}"].alignment = Alignment(horizontal="right")
            row += 1

        # ABC Summary
        if "abc_analysis" in results and "by_product" in results["abc_analysis"]:
            row += 2
            ws[f"A{row}"] = "ABC Classification (Products)"
            ws[f"A{row}"].font = Font(size=14, bold=True)
            ws[f"A{row}"].fill = PatternFill(start_color=ExcelExporter.COLORS["summary"], fill_type="solid")
            ws.merge_cells(f"A{row}:D{row}")

            row += 1
            headers = ["Category", "Count", "% of Items", "Sales %"]
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D9D9D9", fill_type="solid")

            abc_data = results["abc_analysis"]["by_product"]
            for category in ["A", "B", "C"]:
                row += 1
                cat_data = abc_data.get(category, {})
                ws[f"A{row}"] = category
                ws[f"A{row}"].font = Font(bold=True)
                ws[f"A{row}"].fill = PatternFill(
                    start_color=ExcelExporter.COLORS[category], fill_type="solid"
                )
                ws[f"B{row}"] = cat_data.get("count", 0)
                ws[f"C{row}"] = f"{cat_data.get('percentage', 0):.1f}%"
                ws[f"D{row}"] = f"{cat_data.get('sales_pct', 0):.1f}%"

        # Auto-size columns
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 20

    @staticmethod
    def _create_abc_sheet(wb: Workbook, results: Dict, by_type: str, sheet_name: str):
        """Create ABC classification detail sheet"""
        ws = wb.create_sheet(sheet_name)

        abc_data = results["abc_analysis"].get(by_type, {})
        if not abc_data:
            return

        # Headers
        ws["A1"] = f"ABC Classification - {by_type.replace('by_', '').title()}"
        ws["A1"].font = Font(size=14, bold=True)
        ws.merge_cells("A1:D1")

        row = 3
        headers = ["Category", "Count", "% of Total", "Sales", "Sales %"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color=ExcelExporter.COLORS["header"], fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        # Data rows
        for category in ["A", "B", "C"]:
            row += 1
            cat_data = abc_data.get(category, {})

            ws[f"A{row}"] = category
            ws[f"A{row}"].font = Font(bold=True)
            ws[f"A{row}"].fill = PatternFill(
                start_color=ExcelExporter.COLORS[category], fill_type="solid"
            )

            ws[f"B{row}"] = cat_data.get("count", 0)
            ws[f"C{row}"] = f"{cat_data.get('percentage', 0):.2f}%"
            ws[f"D{row}"] = cat_data.get("sales", 0)
            ws[f"D{row}"].number_format = "$#,##0.00"
            ws[f"E{row}"] = f"{cat_data.get('sales_pct', 0):.2f}%"

        # Auto-size columns
        for col in ["A", "B", "C", "D", "E"]:
            ws.column_dimensions[col].width = 15

    @staticmethod
    def _create_top_items_sheet(wb: Workbook, items: List[Dict], sheet_name: str):
        """Create top items (products/clients) sheet"""
        ws = wb.create_sheet(sheet_name)

        if not items:
            return

        # Convert to DataFrame
        df = pd.DataFrame(items)

        # Add title
        ws["A1"] = sheet_name
        ws["A1"].font = Font(size=14, bold=True)
        ws.merge_cells("A1:F1")

        # Write headers
        row = 3
        for col, header in enumerate(df.columns, start=1):
            cell = ws.cell(row=row, column=col, value=header.replace("_", " ").title())
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color=ExcelExporter.COLORS["header"], fill_type="solid")

        # Write data
        for r_idx, row_data in enumerate(df.itertuples(index=False), start=row + 1):
            for c_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)

                # Format numbers
                if c_idx in [2, 4]:  # Sales and avg_price columns
                    cell.number_format = "$#,##0.00"

                # Color by category
                if df.columns[c_idx - 1] == "category":
                    if value in ExcelExporter.COLORS:
                        cell.fill = PatternFill(
                            start_color=ExcelExporter.COLORS[value], fill_type="solid"
                        )

        # Auto-size columns
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column].width = min(max_length + 2, 50)

    @staticmethod
    def _create_discount_sheet(wb: Workbook, discount_data: Dict):
        """Create discount analysis sheet"""
        ws = wb.create_sheet("Discount Analysis")

        # Title
        ws["A1"] = "Discount Analysis"
        ws["A1"].font = Font(size=14, bold=True)
        ws.merge_cells("A1:B1")

        # Summary metrics
        row = 3
        metrics = [
            ("Total Discounts:", f"${discount_data.get('total_discounts', 0):,.2f}"),
            ("Average Discount %:", f"{discount_data.get('avg_discount_pct', 0):.2f}%"),
            ("Rows with Discount:", f"{discount_data.get('rows_with_discount', 0):,}"),
            ("% with Discount:", f"{discount_data.get('percentage_with_discount', 0):.1f}%"),
        ]

        for label, value in metrics:
            ws[f"A{row}"] = label
            ws[f"A{row}"].font = Font(bold=True)
            ws[f"B{row}"] = value
            row += 1

        # Discount by category
        if "discount_by_category" in discount_data:
            row += 2
            ws[f"A{row}"] = "Discounts by ABC Category"
            ws[f"A{row}"].font = Font(bold=True)
            ws.merge_cells(f"A{row}:B{row}")

            row += 1
            for category, amount in discount_data["discount_by_category"].items():
                ws[f"A{row}"] = f"Category {category}"
                ws[f"A{row}"].fill = PatternFill(
                    start_color=ExcelExporter.COLORS.get(category, "FFFFFF"),
                    fill_type="solid",
                )
                ws[f"B{row}"] = amount
                ws[f"B{row}"].number_format = "$#,##0.00"
                row += 1

        # Top discounted products
        if "top_discounted_products" in discount_data:
            top_disc = discount_data["top_discounted_products"]
            if top_disc:
                row += 2
                ws[f"A{row}"] = "Top Discounted Products"
                ws[f"A{row}"].font = Font(size=12, bold=True)
                ws.merge_cells(f"A{row}:D{row}")

                row += 1
                headers = ["Product", "Avg Discount %", "Total Discount $", "Category"]
                for col, header in enumerate(headers, start=1):
                    cell = ws.cell(row=row, column=col, value=header)
                    cell.font = Font(bold=True)

                for item in top_disc[:10]:
                    row += 1
                    ws[f"A{row}"] = item.get("name", "")
                    ws[f"B{row}"] = f"{item.get('avg_discount', 0):.2f}%"
                    ws[f"C{row}"] = item.get("total_discount_amount", 0)
                    ws[f"C{row}"].number_format = "$#,##0.00"
                    ws[f"D{row}"] = item.get("category", "")

        # Auto-size columns
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 20

    @staticmethod
    def _create_margin_sheet(wb: Workbook, margin_data: Dict):
        """Create margin analysis sheet"""
        ws = wb.create_sheet("Margin Analysis")

        # Title
        ws["A1"] = "Margin Analysis"
        ws["A1"].font = Font(size=14, bold=True)
        ws.merge_cells("A1:B1")

        # Summary metrics
        row = 3
        metrics = [
            ("Total Margin:", f"${margin_data.get('total_margin', 0):,.2f}"),
            ("Average Margin %:", f"{margin_data.get('avg_margin_pct', 0):.2f}%"),
        ]

        for label, value in metrics:
            ws[f"A{row}"] = label
            ws[f"A{row}"].font = Font(bold=True)
            ws[f"B{row}"] = value
            row += 1

        # Margin by category
        if "margin_by_category" in margin_data:
            row += 2
            ws[f"A{row}"] = "Margin by ABC Category"
            ws[f"A{row}"].font = Font(bold=True)
            ws.merge_cells(f"A{row}:C{row}")

            row += 1
            headers = ["Category", "Total Margin", "Avg Margin %"]
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = Font(bold=True)

            for category, data in margin_data["margin_by_category"].items():
                row += 1
                ws[f"A{row}"] = f"Category {category}"
                ws[f"A{row}"].fill = PatternFill(
                    start_color=ExcelExporter.COLORS.get(category, "FFFFFF"),
                    fill_type="solid",
                )
                ws[f"B{row}"] = data.get("total_margin", 0)
                ws[f"B{row}"].number_format = "$#,##0.00"
                ws[f"C{row}"] = f"{data.get('avg_margin_pct', 0):.2f}%"

        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 20

    @staticmethod
    def _create_trends_sheet(wb: Workbook, trends: List[Dict]):
        """Create monthly trends sheet"""
        ws = wb.create_sheet("Monthly Trends")

        if not trends:
            return

        # Convert to DataFrame
        df = pd.DataFrame(trends)

        # Title
        ws["A1"] = "Monthly Sales Trends"
        ws["A1"].font = Font(size=14, bold=True)
        ws.merge_cells("A1:E1")

        # Write DataFrame to sheet
        row = 3
        for col, header in enumerate(df.columns, start=1):
            cell = ws.cell(row=row, column=col, value=header.replace("_", " ").title())
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color=ExcelExporter.COLORS["header"], fill_type="solid")

        for r_idx, row_data in enumerate(df.itertuples(index=False), start=row + 1):
            for c_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                if c_idx in [2, 4]:  # Sales and avg_price
                    cell.number_format = "$#,##0.00"
                elif c_idx == 5 and value is not None:  # growth_pct
                    cell.number_format = "0.00%"
                    cell.value = value / 100 if value else 0

        # Auto-size columns
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 15

    @staticmethod
    def _create_insights_sheet(wb: Workbook, insights: List[str]):
        """Create insights sheet"""
        ws = wb.create_sheet("Insights")

        # Title
        ws["A1"] = "Key Insights"
        ws["A1"].font = Font(size=14, bold=True)

        # Write insights
        for idx, insight in enumerate(insights, start=3):
            ws[f"A{idx}"] = f"{idx - 2}. {insight}"
            ws[f"A{idx}"].alignment = Alignment(wrap_text=True, vertical="top")

        ws.column_dimensions["A"].width = 100


class PDFExporter:
    """Export analysis summary to PDF format"""

    @staticmethod
    def export_summary(analysis: Analysis, output_path: str) -> str:
        """
        Export analysis summary to PDF

        Args:
            analysis: Analysis model instance with results
            output_path: Path where to save the PDF file

        Returns:
            Path to the created PDF file

        Raises:
            ValueError: If analysis is not completed or has no results
        """
        if not analysis.is_completed or not analysis.results:
            raise ValueError("Analysis must be completed with results to export")

        logger.info(f"Exporting analysis {analysis.id} to PDF: {output_path}")

        # Create PDF document
        doc = SimpleDocTemplate(output_path, pagesize=letter)
        story = []
        styles = getSampleStyleSheet()

        # Custom styles
        title_style = ParagraphStyle(
            "CustomTitle",
            parent=styles["Heading1"],
            fontSize=24,
            textColor=colors.HexColor("#002060"),
            spaceAfter=30,
            alignment=TA_CENTER,
        )

        heading_style = ParagraphStyle(
            "CustomHeading",
            parent=styles["Heading2"],
            fontSize=16,
            textColor=colors.HexColor("#002060"),
            spaceAfter=12,
        )

        # Title
        story.append(Paragraph("Sales Analysis Report", title_style))
        story.append(Spacer(1, 0.2 * inch))

        # Metadata
        metadata = [
            ["Analysis Name:", analysis.name],
            ["Description:", analysis.description or "N/A"],
            ["Generated:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
            ["Rows Processed:", f"{analysis.row_count:,}"],
        ]

        metadata_table = Table(metadata, colWidths=[2 * inch, 4 * inch])
        metadata_table.setStyle(
            TableStyle(
                [
                    ("FONT", (0, 0), (0, -1), "Helvetica-Bold", 10),
                    ("FONT", (1, 0), (1, -1), "Helvetica", 10),
                    ("ALIGN", (0, 0), (0, -1), "LEFT"),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ]
            )
        )
        story.append(metadata_table)
        story.append(Spacer(1, 0.3 * inch))

        # Summary statistics
        results = analysis.results
        summary = results.get("summary", {})

        story.append(Paragraph("Key Metrics", heading_style))
        metrics_data = [
            ["Metric", "Value"],
            ["Total Sales", f"${summary.get('total_sales', 0):,.2f}"],
            ["Average Sale", f"${summary.get('avg_sale', 0):,.2f}"],
            ["Median Sale", f"${summary.get('median_sale', 0):,.2f}"],
            ["Total Transactions", f"{summary.get('total_rows', 0):,}"],
        ]

        metrics_table = Table(metrics_data, colWidths=[3 * inch, 3 * inch])
        metrics_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#002060")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("FONT", (0, 0), (-1, 0), "Helvetica-Bold", 11),
                    ("FONT", (0, 1), (-1, -1), "Helvetica", 10),
                    ("GRID", (0, 0), (-1, -1), 1, colors.black),
                    ("ALIGN", (1, 0), (1, -1), "RIGHT"),
                ]
            )
        )
        story.append(metrics_table)
        story.append(Spacer(1, 0.3 * inch))

        # ABC Analysis
        if "abc_analysis" in results and "by_product" in results["abc_analysis"]:
            story.append(Paragraph("ABC Classification (Products)", heading_style))

            abc_data = results["abc_analysis"]["by_product"]
            abc_table_data = [
                ["Category", "Count", "% of Items", "Sales %"]
            ]

            for category in ["A", "B", "C"]:
                cat_data = abc_data.get(category, {})
                abc_table_data.append([
                    category,
                    str(cat_data.get("count", 0)),
                    f"{cat_data.get('percentage', 0):.1f}%",
                    f"{cat_data.get('sales_pct', 0):.1f}%",
                ])

            abc_table = Table(abc_table_data, colWidths=[1.5 * inch] * 4)
            abc_table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#002060")),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                        ("FONT", (0, 0), (-1, 0), "Helvetica-Bold", 11),
                        ("FONT", (0, 1), (-1, -1), "Helvetica", 10),
                        ("GRID", (0, 0), (-1, -1), 1, colors.black),
                        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                        ("BACKGROUND", (0, 1), (0, 1), colors.HexColor("#00B050")),
                        ("BACKGROUND", (0, 2), (0, 2), colors.HexColor("#FFC000")),
                        ("BACKGROUND", (0, 3), (0, 3), colors.HexColor("#FF0000")),
                    ]
                )
            )
            story.append(abc_table)
            story.append(Spacer(1, 0.3 * inch))

        # Top 10 Products
        if "top_products" in results:
            story.append(Paragraph("Top 10 Products", heading_style))

            top_products = results["top_products"][:10]
            products_data = [["Rank", "Product", "Sales", "Category"]]

            for idx, product in enumerate(top_products, start=1):
                products_data.append([
                    str(idx),
                    product.get("name", ""),
                    f"${product.get('sales', 0):,.2f}",
                    product.get("category", ""),
                ])

            products_table = Table(
                products_data, colWidths=[0.5 * inch, 3 * inch, 1.5 * inch, 1 * inch]
            )
            products_table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#002060")),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                        ("FONT", (0, 0), (-1, 0), "Helvetica-Bold", 10),
                        ("FONT", (0, 1), (-1, -1), "Helvetica", 9),
                        ("GRID", (0, 0), (-1, -1), 1, colors.black),
                        ("ALIGN", (0, 0), (0, -1), "CENTER"),
                        ("ALIGN", (2, 0), (2, -1), "RIGHT"),
                        ("ALIGN", (3, 0), (3, -1), "CENTER"),
                    ]
                )
            )
            story.append(products_table)
            story.append(Spacer(1, 0.3 * inch))

        # Insights
        if "insights" in results:
            story.append(Paragraph("Key Insights", heading_style))

            for insight in results["insights"]:
                bullet = Paragraph(f"• {insight}", styles["Normal"])
                story.append(bullet)
                story.append(Spacer(1, 0.1 * inch))

        # Build PDF
        doc.build(story)
        logger.info(f"PDF export completed: {output_path}")

        return output_path
